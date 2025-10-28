# -*- coding: utf-8 -*-
# pylint: disable=protected-access
"""The Excel reader to read and chunk Excel files."""
import base64
import hashlib
from typing import Any, Literal, Optional

from ._reader_base import ReaderBase
from ._text_reader import TextReader
from .._document import Document, DocMetadata
from ...message import ImageBlock, Base64Source
from ..._logging import logger


class ExcelReader(ReaderBase):
    """The Excel reader that splits text into chunks by a fixed chunk size."""

    def __init__(
        self,
        chunk_size: int = 512,
        split_by: Literal["char", "sentence", "paragraph"] = "sentence",
        include_sheet_names: bool = True,
        include_cell_coordinates: bool = False,
        include_image: bool = False,
        separate_sheet: bool = False,
    ) -> None:
        """Initialize the Excel reader.

        Args:
            chunk_size (`int`, default to 512):
                The size of each chunk, in number of characters.
            split_by (`Literal["char", "sentence", "paragraph"]`, default to \
            "sentence"):
                The unit to split the text, can be "char", "sentence", or
                "paragraph". The "sentence" option is implemented using the
                "nltk" library, which only supports English text.
            include_sheet_names (`bool`, default to True):
                Whether to include sheet names in the extracted text.
            include_cell_coordinates (`bool`, default to False):
                Whether to include cell coordinates (e.g., A1, B2) in the
                extracted text.
            include_image (`bool`, default to False):
                Whether to include image content in the document. If True,
                images will be extracted and included as base64-encoded images.
            separate_sheet (`bool`, default to False):
                Whether to treat each sheet as a separate document. If True,
                each sheet will be extracted as a separate Document object
                instead of being merged together.
        """
        self._validate_init_params(chunk_size, split_by)

        self.chunk_size = chunk_size
        self.split_by = split_by
        self.include_sheet_names = include_sheet_names
        self.include_cell_coordinates = include_cell_coordinates
        self.include_image = include_image
        self.separate_sheet = separate_sheet

        # To avoid code duplication, we use TextReader to do the chunking.
        self._text_reader = TextReader(self.chunk_size, self.split_by)

        # Cache for imports (none needed for direct imports)

    def _validate_init_params(self, chunk_size: int, split_by: str) -> None:
        """Validate initialization parameters.

        Args:
            chunk_size (`int`):
                The chunk size to validate.
            split_by (`str`):
                The split mode to validate.
        """
        if chunk_size <= 0:
            raise ValueError(
                f"The chunk_size must be positive, got {chunk_size}",
            )

        if split_by not in ["char", "sentence", "paragraph"]:
            raise ValueError(
                "The split_by must be one of 'char', 'sentence' or "
                f"'paragraph', got {split_by}",
            )

    async def __call__(
        self,
        excel_path: str,
    ) -> list[Document]:
        """Read an Excel file, split it into chunks, and return a list of
        Document objects.

        Args:
            excel_path (`str`):
                The input Excel file path (.xlsx or .xls file).

        Returns:
            `list[Document]`:
                A list of Document objects, where the metadata contains the
                chunked text, doc id and chunk id.
        """
        # Generate document ID
        doc_id = self.get_doc_id(excel_path)

        # Initialize variables
        excel_file = None
        workbook = None

        try:
            import pandas as pd
        except ImportError as e:
            raise ImportError(
                "Please install pandas to use the Excel reader. "
                "You can install it by `pip install pandas`.",
            ) from e

        try:
            excel_file = pd.ExcelFile(excel_path)

            # Load workbook if images are needed
            if self.include_image:
                try:
                    from openpyxl import load_workbook

                    workbook = load_workbook(excel_path)
                except ImportError:
                    logger.warning(
                        "openpyxl not available, image extraction disabled",
                    )
                    workbook = None

            # Process sheets (images will be extracted per-sheet to
            # maintain order)
            if self.separate_sheet:
                return await self._process_sheets_separately(
                    excel_file,
                    doc_id,
                    workbook,
                )
            else:
                return await self._process_sheets_merged(
                    excel_file,
                    doc_id,
                    workbook,
                )

        except (
            pd.errors.EmptyDataError,
            pd.errors.ParserError,
            FileNotFoundError,
            PermissionError,
        ) as e:
            raise ValueError(
                f"Failed to read Excel file {excel_path}: {e}",
            ) from e
        finally:
            # Ensure all resources are closed
            if workbook is not None:
                workbook.close()
            if excel_file is not None:
                excel_file.close()

    def _extract_sheet_images_from_workbook(
        self,
        workbook: Any,
        sheet_name: str,
        doc_id: str,
    ) -> list[Document]:
        """Extract images from a specific sheet using already loaded workbook.

        Args:
            workbook (`Any`):
                The openpyxl workbook object.
            sheet_name (`str`):
                The name of the sheet.
            doc_id (`str`):
                The document ID.

        Returns:
            `list[Document]`:
                A list of Document objects containing images.
        """
        try:
            ws = workbook[sheet_name]
            return self._extract_sheet_images(ws, doc_id, sheet_name)
        except Exception as e:
            logger.warning(
                "Failed to extract images from sheet '%s': %s",
                sheet_name,
                e,
            )
            return []

    async def _process_sheets_merged(
        self,
        excel_file: Any,
        doc_id: str,
        workbook: Any = None,
    ) -> list[Document]:
        """Process all sheets as a single merged document.

        Args:
            excel_file (`Any`):
                The pandas ExcelFile object.
            doc_id (`str`):
                The document ID.
            workbook (`Any`, optional):
                The openpyxl workbook if available.

        Returns:
            `list[Document]`:
                A list of Document objects from all sheets merged together.
        """
        # Collect all text from all sheets
        all_text_parts = []

        # Collect all images from all sheets
        all_images = []

        for sheet_name in excel_file.sheet_names:
            # Extract images for this sheet if requested
            if self.include_image and workbook:
                sheet_images = self._extract_sheet_images_from_workbook(
                    workbook,
                    sheet_name,
                    doc_id,
                )
                all_images.extend(sheet_images)

            # Process text content for this sheet
            sheet_text = self._process_single_sheet_from_file(
                excel_file,
                sheet_name,
            )

            if sheet_text:
                all_text_parts.append(sheet_text)

        # Merge all text and create documents
        merged_text = "\n\n".join(all_text_parts)

        # Add images first (if any), then merged text
        all_docs = []
        if all_images:
            all_docs.extend(all_images)

        if merged_text:
            text_docs = await self._create_documents_from_text(
                merged_text,
                doc_id,
            )
            all_docs.extend(text_docs)

        return all_docs

    async def _process_sheets_separately(
        self,
        excel_file: Any,
        doc_id: str,
        workbook: Any = None,
    ) -> list[Document]:
        """Process each sheet as separate documents.

        Args:
            excel_file (`Any`):
                The pandas ExcelFile object.
            doc_id (`str`):
                The document ID.
            workbook (`Any`, optional):
                The openpyxl workbook if available.

        Returns:
            `list[Document]`:
                A list of Document objects with each sheet processed
                separately.
        """
        all_docs = []

        for sheet_name in excel_file.sheet_names:
            # Extract images for this sheet if requested
            sheet_images = []
            if self.include_image and workbook is not None:
                sheet_images = self._extract_sheet_images_from_workbook(
                    workbook,
                    sheet_name,
                    doc_id,
                )

            # Process text content for this sheet (use already loaded
            # excel_file)
            sheet_text = self._process_single_sheet_from_file(
                excel_file,
                sheet_name,
            )

            # Add images first (if any), then text for this sheet
            if sheet_images:
                all_docs.extend(sheet_images)

            if sheet_text:
                sheet_docs = await self._create_documents_from_text(
                    sheet_text,
                    doc_id,
                )
                all_docs.extend(sheet_docs)

        return all_docs

    def _process_single_sheet_from_file(
        self,
        excel_file: Any,
        sheet_name: str,
    ) -> Optional[str]:
        """Process a single sheet from already loaded ExcelFile and return
        its text content.

        Args:
            excel_file (`Any`):
                The pandas ExcelFile object.
            sheet_name (`str`):
                The name of the sheet to process.

        Returns:
            `Optional[str]`:
                The text content of the sheet, or None if empty.
        """
        try:
            df = excel_file.parse(sheet_name=sheet_name)

            if df.empty:
                return None

            sheet_texts = []

            # Add sheet name if requested
            if self.include_sheet_names:
                sheet_texts.append(f"Sheet: {sheet_name}")

            # Process rows
            row_texts = self._process_dataframe_rows(df)
            sheet_texts.extend(row_texts)

            return "\n".join(sheet_texts) if sheet_texts else None

        except Exception as e:
            logger.warning("Failed to process sheet '%s': %s", sheet_name, e)
            return None

    def _process_dataframe_rows(self, df: Any) -> list[str]:
        """Process DataFrame rows and return list of row texts.

        Args:
            df (`Any`):
                The pandas DataFrame to process.

        Returns:
            `list[str]`:
                List of row texts.
        """
        row_texts = []

        for row_idx, row in df.iterrows():
            cell_texts = []

            for col_idx, (_, cell_val) in enumerate(row.items()):
                cell_text = self._process_cell_value(
                    cell_val,
                    row_idx,
                    col_idx,
                )
                if cell_text:
                    cell_texts.append(cell_text)

            if cell_texts:
                row_text = " | ".join(cell_texts)
                row_texts.append(row_text)

        return row_texts

    def _process_cell_value(
        self,
        cell_val: Any,
        row_idx: int,
        col_idx: int,
    ) -> Optional[str]:
        """Process individual cell value.

        Args:
            cell_val (`Any`):
                The cell value from the DataFrame.
            row_idx (`int`):
                The row index.
            col_idx (`int`):
                The column index.

        Returns:
            `Optional[str]`:
                The processed cell text, or None if empty.
        """
        # Skip NaN values
        import pandas as pd

        if pd.isna(cell_val):
            return None

        # Convert to string and clean
        cell_text = str(cell_val).strip()
        if not cell_text:
            return None

        # Add cell coordinates if requested
        if self.include_cell_coordinates:
            col_letter = self._get_column_letter(col_idx)
            cell_ref = f"{col_letter}{row_idx + 1}"
            cell_text = f"{cell_ref}: {cell_text}"

        # Clean text
        return self._clean_text(cell_text)

    def _clean_text(self, text: str) -> str:
        """Clean and normalize text content.

        Args:
            text (`str`):
                The text to clean.

        Returns:
            `str`:
                The cleaned text.
        """
        # Normalize line endings
        text = text.replace("\r\n", "\n").replace("\r", "\n")

        # Remove control characters except newlines and tabs
        cleaned_chars = []
        for char in text:
            char_code = ord(char)
            if char_code >= 32 or char in "\n\t":
                cleaned_chars.append(char)

        return "".join(cleaned_chars)

    def _get_column_letter(self, col_idx: int) -> str:
        """Convert column index to Excel column letter.

        Args:
            col_idx (`int`):
                Zero-based column index.

        Returns:
            `str`:
                Excel column letter (A, B, C, ..., AA, AB, ...).
        """
        result = ""
        while col_idx >= 0:
            result = chr(65 + (col_idx % 26)) + result
            col_idx = col_idx // 26 - 1
        return result

    async def _create_documents_from_text(
        self,
        text: str,
        doc_id: str,
    ) -> list[Document]:
        """Create Document objects from text using TextReader.

        Args:
            text (`str`):
                The text to convert to documents.
            doc_id (`str`):
                The document ID.

        Returns:
            `list[Document]`:
                A list of Document objects.
        """
        try:
            docs = await self._text_reader(text)
            # TextReader already returns properly formatted documents,
            # just update doc_id
            for doc in docs:
                doc.id = doc_id
                doc.metadata.doc_id = doc_id
            return docs
        except Exception as e:
            logger.error("Failed to create documents from text: %s", e)
            return []

    def _extract_sheet_images(
        self,
        worksheet: Any,
        doc_id: str,
        sheet_name: str,
    ) -> list[Document]:
        """Extract images from a single worksheet.

        Args:
            worksheet (`Any`):
                The openpyxl worksheet object.
            doc_id (`str`):
                The document ID.
            sheet_name (`str`):
                The name of the sheet.

        Returns:
            `list[Document]`:
                A list of Document objects containing images.
        """
        images = []

        if not (hasattr(worksheet, "_images") and worksheet._images):
            return images

        for idx, img in enumerate(worksheet._images):
            try:
                image_doc = self._create_image_document(
                    img,
                    doc_id,
                    sheet_name,
                    idx,
                )
                if image_doc:
                    images.append(image_doc)
            except Exception as e:
                logger.warning(
                    "Failed to process image %d from sheet '%s': %s",
                    idx,
                    sheet_name,
                    e,
                )
                continue

        return images

    def _create_image_document(
        self,
        img: Any,
        doc_id: str,
        sheet_name: str,
        img_idx: int,
    ) -> Optional[Document]:
        """Create Document object from image.

        Args:
            img (`Any`):
                The image object from openpyxl.
            doc_id (`str`):
                The document ID.
            sheet_name (`str`):
                The name of the sheet containing the image.
            img_idx (`int`):
                The index of the image in the sheet.

        Returns:
            `Optional[Document]`:
                A Document object containing the image, or None on error.
        """
        try:
            # Get image data
            img_data = img._data()

            # Determine media type
            media_type = self._get_media_type_from_data(img_data)

            # Convert to base64
            base64_data = base64.b64encode(img_data).decode("utf-8")

            # Create ImageBlock
            image_block = ImageBlock(
                type="image",
                source=Base64Source(
                    type="base64",
                    media_type=media_type,
                    data=base64_data,
                ),
            )

            # Create Document - each image is an independent document
            metadata = DocMetadata(
                content=image_block,
                doc_id=doc_id,
                chunk_id=0,
                total_chunks=1,
            )
            # Add additional metadata
            metadata["sheet_name"] = sheet_name
            metadata["img_idx"] = img_idx

            return Document(metadata=metadata)
        except Exception as e:
            logger.error("Failed to create image document: %s", e)
            return None

    def _get_media_type_from_data(self, data: bytes) -> str:
        """Determine media type from image data.

        Args:
            data (`bytes`):
                The raw image data.

        Returns:
            `str`:
                The MIME type of the image (e.g., "image/png", "image/jpeg").
        """
        # Image signature mapping
        signatures = {
            b"\x89PNG\r\n\x1a\n": "image/png",
            b"\xff\xd8": "image/jpeg",
            b"GIF87a": "image/gif",
            b"GIF89a": "image/gif",
            b"BM": "image/bmp",
        }

        # Check signatures
        for signature, media_type in signatures.items():
            if data.startswith(signature):
                return media_type

        # Check WebP (RIFF at start + WEBP at offset 8)
        if len(data) > 12 and data[:4] == b"RIFF" and data[8:12] == b"WEBP":
            return "image/webp"

        # Default to JPEG
        return "image/jpeg"

    def get_doc_id(self, excel_path: str) -> str:
        """Generate unique document ID from file path.

        Args:
            excel_path (`str`):
                The path to the Excel file.

        Returns:
            `str`:
                The document ID (SHA256 hash of the file path).
        """
        return hashlib.sha256(excel_path.encode("utf-8")).hexdigest()
