# -*- coding: utf-8 -*-
# pylint: disable=protected-access
"""Excel (.xlsx / .xls) file parser.

Reads an Excel workbook sheet-by-sheet and emits :class:`Section`
objects.  Each sheet's tabular content is rendered as a single text
section (in Markdown or JSON format); embedded images (when enabled)
are emitted as their own :class:`DataBlock` sections.

Supported options: ``include_sheet_names``, ``include_cell_coordinates``,
``include_image``, ``separate_sheet``, ``table_format``.  Chunking is
**not** done here — long text stays intact inside a section and is split
downstream by a :class:`~agentscope.rag.ChunkerBase`.
"""
import base64
import io
import json
from typing import Any, Literal

from ..._logging import logger
from ...message import Base64Source, DataBlock, TextBlock
from .._document import Section
from ._base import ParserBase
from ._utils import _guess_image_media_type


def _get_excel_column_name(col_index: int) -> str:
    """Convert a 0-based column index to an Excel column name.

    Args:
        col_index (`int`):
            Zero-based column index (0 → ``"A"``, 25 → ``"Z"``,
            26 → ``"AA"``, etc.).

    Returns:
        `str`:
            The corresponding Excel column letter(s).
    """
    result = ""
    col_index += 1
    while col_index > 0:
        col_index -= 1
        result = chr(ord("A") + col_index % 26) + result
        col_index //= 26
    return result


def _extract_table_data(df: Any) -> list[list[str]]:
    """Extract table data from a pandas DataFrame.

    NaN values are converted to empty strings, and Windows-style line
    breaks (``\\r\\n``) are normalised to ``\\n``.

    Args:
        df (`pandas.DataFrame`):
            The DataFrame to extract data from.

    Returns:
        `list[list[str]]`:
            A 2-D list where the first row is the column header and
            subsequent rows contain the cell values as strings.
    """
    import pandas as pd

    header = [str(col).strip() for col in df.columns]
    rows: list[list[str]] = [header]
    for _, row in df.iterrows():
        cells: list[str] = []
        for val in row:
            if pd.isna(val):
                cells.append("")
            else:
                text = str(val).strip()
                text = text.replace("\r\n", "\n").replace("\r", "\n")
                cells.append(text)
        rows.append(cells)
    return rows


def _extract_images_from_worksheet(
    worksheet: Any,
    filename: str,
) -> list[tuple[int, DataBlock]]:
    """Extract images from an openpyxl worksheet with their row positions.

    Args:
        worksheet (`openpyxl.worksheet.worksheet.Worksheet`):
            The openpyxl worksheet to scan for embedded images.
        filename (`str`):
            Source filename, stored in each :class:`DataBlock`'s
            ``name`` field.

    Returns:
        `list[tuple[int, DataBlock]]`:
            A list of ``(row_index, DataBlock)`` tuples, where
            ``row_index`` is 0-based.  Images whose anchor cannot
            be determined default to row 0.
    """
    images: list[tuple[int, DataBlock]] = []

    if not (hasattr(worksheet, "_images") and worksheet._images):
        return images

    for img in worksheet._images:
        try:
            row_index = 0
            if hasattr(img, "anchor") and hasattr(img.anchor, "_from"):
                row_index = img.anchor._from.row

            img_data = img._data()
            media_type = _guess_image_media_type(img_data)
            b64_data = base64.b64encode(img_data).decode("utf-8")

            images.append(
                (
                    row_index,
                    DataBlock(
                        source=Base64Source(
                            media_type=media_type,
                            data=b64_data,
                        ),
                        name=filename,
                    ),
                ),
            )
        except Exception as e:
            logger.warning("Failed to extract image from worksheet: %s", e)

    return images


class ExcelParser(ParserBase):
    """Parser for Excel ``.xlsx`` / ``.xls`` files.

    Each sheet is scanned for tabular data and (optionally) images.
    Tables are rendered as Markdown pipe-tables or JSON arrays; images
    are emitted as standalone :class:`DataBlock` sections.

    When ``separate_sheet=True`` each sheet becomes a batch of
    sections that never intermix with other sheets, making it possible
    for a downstream chunker to keep sheets apart.

    .. note:: The table content can be extracted in Markdown or JSON format.

        **Markdown format example** (``include_cell_coordinates=False``):

        .. code-block:: text

            | Name  | Age | City     |
            |-------|-----|----------|
            | Alice | 25  | New York |

        **JSON format example** (``include_cell_coordinates=True``):

        .. code-block:: json

            {"A1": "Name", "B1": "Age", "C1": "City"}
            {"A2": "Alice", "B2": "25", "C2": "New York"}
    """

    supported_media_types: list[str] = [
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        "application/vnd.ms-excel",
    ]

    @classmethod
    def supported_extensions(cls) -> list[str]:
        """Return ``[".xlsx", ".xls"]``."""
        return [".xls", ".xlsx"]

    def __init__(
        self,
        include_sheet_names: bool = True,
        include_cell_coordinates: bool = False,
        include_image: bool = False,
        separate_sheet: bool = False,
        table_format: Literal["markdown", "json"] = "markdown",
    ) -> None:
        """Initialize the Excel parser.

        Args:
            include_sheet_names (`bool`, defaults to ``True``):
                Whether to prepend each sheet's name as a header.
            include_cell_coordinates (`bool`, defaults to ``False``):
                Whether to include cell coordinates (e.g. ``[A1]``) in
                the rendered table.
            include_image (`bool`, defaults to ``False``):
                Whether to extract and include embedded images.
                Requires ``openpyxl``.
            separate_sheet (`bool`, defaults to ``False``):
                If ``True``, images and table text from different
                sheets are never combined into the same section.
            table_format (`Literal["markdown", "json"]``, defaults to
                ``"markdown"``):
                How to render tables.

        Raises:
            `ValueError`: If ``table_format`` is not ``"markdown"``
                or ``"json"``.
        """
        if table_format not in ("markdown", "json"):
            raise ValueError(
                "The table_format must be one of 'markdown' or 'json', "
                f"got {table_format!r}.",
            )
        self.include_sheet_names = include_sheet_names
        self.include_cell_coordinates = include_cell_coordinates
        self.include_image = include_image
        self.separate_sheet = separate_sheet
        self.table_format = table_format

    async def parse(
        self,
        file: bytes | str,
        filename: str,
    ) -> list[Section]:
        """Parse an Excel file into a list of :class:`Section` objects.

        Args:
            file (`bytes | str`):
                Either the raw Excel bytes, or a filesystem path to
                the Excel file.
            filename (`str`):
                The source filename, copied into each Section's
                :attr:`Section.source`.

        Returns:
            `list[Section]`:
                Sections in sheet order.  When ``separate_sheet=True``,
                text sections carry ``{"sheet": "<name>"}`` metadata
                and image sections add
                ``{"media_type": "image/..."}``.  When
                ``separate_sheet=False`` (default), all text is merged
                into a single section with ``metadata={}``.

        Raises:
            `FileNotFoundError`: If ``file`` is a ``str`` pointing to
                a path that does not exist.
            `ImportError`: If :mod:`pandas` is not installed.
            `ValueError`: If the bytes cannot be parsed.
        """
        try:
            import pandas as pd
        except ImportError as e:
            raise ImportError(
                "Please install pandas to use the Excel parser. "
                "You can install it by `pip install pandas` (or "
                "`pip install agentscope[rag]`).",
            ) from e

        if isinstance(file, str):
            excel_file = pd.ExcelFile(file)
        else:
            excel_file = pd.ExcelFile(io.BytesIO(file))

        workbook = None
        try:
            if self.include_image:
                try:
                    from openpyxl import load_workbook

                    if isinstance(file, str):
                        workbook = load_workbook(file)
                    else:
                        workbook = load_workbook(io.BytesIO(file))
                except ImportError:
                    logger.warning(
                        "openpyxl not available, image extraction disabled.",
                    )
                except Exception as e:
                    logger.warning("Failed to load workbook for images: %s", e)

            sections: list[Section] = []

            for sheet_name in excel_file.sheet_names:
                sheet_sections = self._parse_sheet(
                    excel_file,
                    sheet_name,
                    filename,
                    workbook,
                )
                sections.extend(sheet_sections)

            if not self.separate_sheet:
                sections = self._merge_text_sections(sections, filename)

            return sections
        finally:
            if workbook is not None:
                workbook.close()
            excel_file.close()

    # ------------------------------------------------------------------
    # Sheet-level parsing
    # ------------------------------------------------------------------

    def _parse_sheet(
        self,
        excel_file: Any,
        sheet_name: str,
        filename: str,
        workbook: Any = None,
    ) -> list[Section]:
        """Parse a single sheet and return its sections."""
        sheet_sections: list[Section] = []

        try:
            df = excel_file.parse(sheet_name=sheet_name)
        except Exception as e:
            logger.warning("Failed to parse sheet '%s': %s", sheet_name, e)
            return sheet_sections

        if df.empty:
            return sheet_sections

        table_data = _extract_table_data(df)

        if self.table_format == "markdown":
            table_text = self._table_to_markdown(table_data, sheet_name)
        else:
            table_text = self._table_to_json(table_data, sheet_name)

        if table_text:
            sheet_sections.append(
                Section(
                    content=TextBlock(text=table_text),
                    source=filename,
                    metadata={"sheet": sheet_name},
                ),
            )

        if self.include_image and workbook is not None:
            try:
                worksheet = workbook[sheet_name]
                images = _extract_images_from_worksheet(
                    worksheet,
                    filename,
                )
                images.sort(key=lambda x: x[0])

                for _row, block in images:
                    sheet_sections.append(
                        Section(
                            content=block,
                            source=filename,
                            metadata={
                                "sheet": sheet_name,
                                "media_type": block.source.media_type,
                            },
                        ),
                    )
            except Exception as e:
                logger.warning(
                    "Failed to extract images from sheet '%s': %s",
                    sheet_name,
                    e,
                )

        return sheet_sections

    # ------------------------------------------------------------------
    # Cross-sheet merging
    # ------------------------------------------------------------------

    @staticmethod
    def _merge_text_sections(
        sections: list[Section],
        filename: str,
    ) -> list[Section]:
        """Merge all text sections into a single section.

        When ``separate_sheet=False`` (default), text from all sheets
        is concatenated into one :class:`Section` so a downstream
        chunker may create chunks that span sheet boundaries.  Image
        sections are kept unchanged and appended after the merged text.
        """
        text_parts: list[str] = []
        non_text: list[Section] = []
        for section in sections:
            if isinstance(section.content, TextBlock):
                text_parts.append(section.content.text)
            else:
                non_text.append(section)

        result: list[Section] = []
        if text_parts:
            result.append(
                Section(
                    content=TextBlock(text="\n".join(text_parts)),
                    source=filename,
                    metadata={},
                ),
            )
        result.extend(non_text)
        return result

    # ------------------------------------------------------------------
    # Table rendering
    # ------------------------------------------------------------------

    def _table_to_markdown(
        self,
        table_data: list[list[str]],
        sheet_name: str,
    ) -> str:
        """Render table data as Markdown with optional sheet header and
        cell coordinates."""
        if not table_data or not table_data[0]:
            return ""

        lines: list[str] = []

        if self.include_sheet_names:
            lines.append(f"Sheet: {sheet_name}")

        num_cols = len(table_data[0])

        def _fmt(cell: str, row_idx: int, col_idx: int) -> str:
            escaped = cell.replace("|", "\\|")
            if self.include_cell_coordinates:
                coord = f"{_get_excel_column_name(col_idx)}{row_idx + 1}"
                return f"[{coord}] {escaped}"
            return escaped

        header = [_fmt(c, 0, ci) for ci, c in enumerate(table_data[0])]
        lines.append("| " + " | ".join(header) + " |")
        lines.append("| " + " | ".join(["---"] * num_cols) + " |")

        for ri, row in enumerate(table_data[1:], start=1):
            padded = list(row) + [""] * max(0, num_cols - len(row))
            cells = [_fmt(c, ri, ci) for ci, c in enumerate(padded[:num_cols])]
            lines.append("| " + " | ".join(cells) + " |")

        return "\n".join(lines) + "\n"

    def _table_to_json(
        self,
        table_data: list[list[str]],
        sheet_name: str,
    ) -> str:
        """Render table data as JSON with optional sheet header and cell
        coordinates."""
        parts: list[str] = []

        if self.include_sheet_names:
            parts.append(f"Sheet: {sheet_name}")

        parts.append(
            "<system-info>A table loaded as a JSON array:</system-info>",
        )

        for ri, row in enumerate(table_data):
            if self.include_cell_coordinates:
                row_dict = {
                    f"{_get_excel_column_name(ci)}{ri + 1}": cell
                    for ci, cell in enumerate(row)
                }
                parts.append(json.dumps(row_dict, ensure_ascii=False))
            else:
                parts.append(json.dumps(row, ensure_ascii=False))

        return "\n".join(parts)
